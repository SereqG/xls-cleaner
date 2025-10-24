from langchain.agents import AgentExecutor, create_openai_tools_agent
from langchain_openai import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain.prompts import ChatPromptTemplate, MessagesPlaceholder
from langchain.tools import Tool
from typing import Dict, List, Any
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
from config import Config


class ExcelModificationTool:
    """Tool for modifying Excel files based on user instructions."""
    
    def __init__(self, file_path: str, sheet_name: str):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.df = None
        self._load_workbook()
    
    def _load_workbook(self):
        """Load the Excel workbook and convert sheet to DataFrame."""
        self.workbook = load_workbook(self.file_path)
        if self.sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{self.sheet_name}' not found in workbook")
        
        # Load sheet data into pandas DataFrame
        self.df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
    
    def get_data_info(self, query: str) -> str:
        """Get information about the data."""
        try:
            info = f"Sheet: {self.sheet_name}\n"
            info += f"Rows: {len(self.df)}\n"
            info += f"Columns: {', '.join(self.df.columns.tolist())}\n"
            info += f"Column types:\n{self.df.dtypes.to_string()}\n"
            info += f"\nFirst 3 rows:\n{self.df.head(3).to_string()}"
            return info
        except Exception as e:
            return f"Error getting data info: {str(e)}"
    
    def modify_data(self, instructions: str) -> str:
        """
        Modify the Excel data based on instructions.
        
        Examples of supported operations:
        - Remove rows where column X is empty
        - Fill empty cells in column Y with value Z
        - Convert column A to uppercase
        - Remove duplicate rows
        - Filter rows where column B > value
        """
        try:
            original_rows = len(self.df)
            
            # Parse and execute instructions
            instructions_lower = instructions.lower()
            
            if "remove empty" in instructions_lower or "delete empty" in instructions_lower:
                # Extract column name and remove empty rows
                for col in self.df.columns:
                    if col.lower() in instructions_lower:
                        before = len(self.df)
                        self.df = self.df[self.df[col].notna()]
                        after = len(self.df)
                        return f"Removed {before - after} rows with empty values in column '{col}'"
            
            elif "fill empty" in instructions_lower or "replace empty" in instructions_lower:
                # Fill empty cells with specified value
                for col in self.df.columns:
                    if col.lower() in instructions_lower:
                        # Try to extract fill value
                        words = instructions.split()
                        if "with" in words:
                            idx = words.index("with")
                            if idx + 1 < len(words):
                                fill_value = words[idx + 1].strip('"\'')
                                count = self.df[col].isna().sum()
                                self.df[col].fillna(fill_value, inplace=True)
                                return f"Filled {count} empty cells in column '{col}' with '{fill_value}'"
            
            elif "uppercase" in instructions_lower or "upper case" in instructions_lower:
                # Convert column to uppercase
                for col in self.df.columns:
                    if col.lower() in instructions_lower and self.df[col].dtype == 'object':
                        self.df[col] = self.df[col].astype(str).str.upper()
                        return f"Converted column '{col}' to uppercase"
            
            elif "lowercase" in instructions_lower or "lower case" in instructions_lower:
                # Convert column to lowercase
                for col in self.df.columns:
                    if col.lower() in instructions_lower and self.df[col].dtype == 'object':
                        self.df[col] = self.df[col].astype(str).str.lower()
                        return f"Converted column '{col}' to lowercase"
            
            elif "remove duplicate" in instructions_lower or "delete duplicate" in instructions_lower:
                before = len(self.df)
                self.df = self.df.drop_duplicates()
                after = len(self.df)
                return f"Removed {before - after} duplicate rows"
            
            elif "sort" in instructions_lower:
                # Sort by column
                for col in self.df.columns:
                    if col.lower() in instructions_lower:
                        ascending = "descending" not in instructions_lower and "desc" not in instructions_lower
                        self.df = self.df.sort_values(by=col, ascending=ascending)
                        order = "ascending" if ascending else "descending"
                        return f"Sorted data by column '{col}' in {order} order"
            
            return "I couldn't understand the modification instruction. Please be more specific about which column to modify and what operation to perform."
            
        except Exception as e:
            return f"Error modifying data: {str(e)}"
    
    def save_workbook(self, output_path: str) -> bool:
        """Save the modified workbook."""
        try:
            # Write DataFrame back to Excel
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                self.df.to_excel(writer, sheet_name=self.sheet_name, index=False)
            return True
        except Exception as e:
            print(f"Error saving workbook: {str(e)}")
            return False
    
    def get_preview(self, num_rows: int = 5) -> str:
        """Get a preview of the current data."""
        try:
            preview = self.df.head(num_rows).to_string()
            return f"Preview (first {num_rows} rows):\n{preview}"
        except Exception as e:
            return f"Error getting preview: {str(e)}"


class AIAgentService:
    """Service for managing AI agents with memory for Excel file processing."""
    
    def __init__(self):
        self.agents: Dict[str, AgentExecutor] = {}
        self.memories: Dict[str, ConversationBufferMemory] = {}
        self.tools: Dict[str, ExcelModificationTool] = {}
    
    def _is_excel_related(self, message: str) -> bool:
        """Check if the message is Excel-related."""
        excel_keywords = [
            'excel', 'spreadsheet', 'column', 'row', 'cell', 'data',
            'sheet', 'table', 'remove', 'delete', 'fill', 'modify',
            'change', 'update', 'filter', 'sort', 'clean', 'format',
            'empty', 'duplicate', 'value', 'sum', 'average', 'count'
        ]
        message_lower = message.lower()
        return any(keyword in message_lower for keyword in excel_keywords)
    
    def create_agent(self, session_id: str, file_path: str, sheet_name: str) -> AgentExecutor:
        """Create a new agent for a session with Excel tools."""
        
        # Create Excel modification tool
        excel_tool = ExcelModificationTool(file_path, sheet_name)
        self.tools[session_id] = excel_tool
        
        # Define tools for the agent
        tools = [
            Tool(
                name="get_data_info",
                func=excel_tool.get_data_info,
                description="Get information about the Excel data including columns, types, and preview"
            ),
            Tool(
                name="modify_data",
                func=excel_tool.modify_data,
                description="Modify the Excel data based on user instructions. Supports operations like removing empty rows, filling empty cells, converting to uppercase/lowercase, removing duplicates, sorting, etc."
            ),
            Tool(
                name="get_preview",
                func=excel_tool.get_preview,
                description="Get a preview of the current data state"
            )
        ]
        
        # Create memory
        memory = ConversationBufferMemory(
            memory_key="chat_history",
            return_messages=True
        )
        self.memories[session_id] = memory
        
        # Create prompt
        prompt = ChatPromptTemplate.from_messages([
            ("system", """You are an Excel data cleaning assistant. You help users clean and modify their Excel spreadsheets.
            
You have access to tools to view and modify Excel data. When a user asks you to make changes:
1. First use get_data_info to understand the data structure
2. Then use modify_data to make the requested changes
3. Finally use get_preview to show the user what changed

Be helpful and suggest improvements when appropriate. Always explain what you're doing.

IMPORTANT: Only respond to Excel-related questions. If a user asks about anything not related to Excel or data processing, respond with: "I can't respond to non Excel-related prompts"."""),
            MessagesPlaceholder(variable_name="chat_history"),
            ("human", "{input}"),
            MessagesPlaceholder(variable_name="agent_scratchpad"),
        ])
        
        # Create LLM
        llm = ChatOpenAI(
            temperature=0,
            model="gpt-3.5-turbo",
            openai_api_key=Config.OPENAI_API_KEY
        )
        
        # Create agent
        agent = create_openai_tools_agent(llm, tools, prompt)
        agent_executor = AgentExecutor(
            agent=agent,
            tools=tools,
            memory=memory,
            verbose=True,
            handle_parsing_errors=True
        )
        
        self.agents[session_id] = agent_executor
        return agent_executor
    
    def get_agent(self, session_id: str) -> AgentExecutor:
        """Get an existing agent by session ID."""
        return self.agents.get(session_id)
    
    def chat(self, session_id: str, message: str) -> str:
        """Send a message to the agent and get a response."""
        # Check if message is Excel-related
        if not self._is_excel_related(message):
            return "I can't respond to non Excel-related prompts"
        
        agent = self.get_agent(session_id)
        if not agent:
            return "Agent not found. Please start a new session."
        
        try:
            response = agent.invoke({"input": message})
            return response.get("output", "I couldn't process that request.")
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error in AI chat: {str(e)}", exc_info=True)
            return "I encountered an error processing your request. Please try again."
    
    def save_modified_file(self, session_id: str, output_path: str) -> bool:
        """Save the modified Excel file."""
        tool = self.tools.get(session_id)
        if not tool:
            return False
        return tool.save_workbook(output_path)
    
    def get_preview(self, session_id: str, num_rows: int = 5) -> str:
        """Get a preview of the current data state."""
        tool = self.tools.get(session_id)
        if not tool:
            return "No data available"
        return tool.get_preview(num_rows)
    
    def cleanup_session(self, session_id: str):
        """Clean up agent resources for a session."""
        if session_id in self.agents:
            del self.agents[session_id]
        if session_id in self.memories:
            del self.memories[session_id]
        if session_id in self.tools:
            del self.tools[session_id]


# Global service instance
ai_service = AIAgentService()
